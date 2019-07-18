import openpyxl

xl = openpyxl.load_workbook("D:\\vip_dataTable.xlsx")

print(type(xl))
print(xl.sheetnames)
print(xl.active.title)
print("Maximum Rows:", xl.active.max_row)
print("Maximum Columns:", xl.active.max_column)

# sheet = xl.get_sheet_by_name('Sheet1')
print(xl.active.columns[1])

# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, ":", cellObj.value)

for cellObj in xl.active.columns('1'):
        print(cellObj.value)
